#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import openpyxl


#单词
class Words:
                    
  #表記  読み  声调  词性  词义  文節（詞組）  意味
  def __init__(self, spell, pron, accent, word_class , meanings , sentence , translation):
    self.spell = spell
    self.pron = pron
    self.accent = accent
    self.word_class = word_class
    self.meanings = meanings
    self.sentence = sentence
    self.translation = translation

  def __eq__(self, other):
    if not isinstance(other, Words):
      return NotImplemented
    else:
      # string lists of all method names and properties of each of these objects
      prop_names1 = list(self.__dict__)
      prop_names2 = list(other.__dict__)

      n = len(prop_names1)  # number of properties
      for i in range(n):
        if getattr(self, prop_names1[i]) != getattr(other, prop_names2[i]):
          return False

      return True

  def __hash__(self):
    # necessary for instances to behave sanely in dicts and sets.
    return hash((self.foo, self.bar))
  
  def __str__(self):
    return "%s | %s | %s | %s | %s | %s | %s " %(self.spell, self.pron, self.accent, self.word_class , self.meanings , self.sentence , self.translation)

  #是否要更新  
  def need_fix(self):
    return not (self.spell and self.pron and self.accent and self.word_class and self.meanings and self.sentence and self.translation)


  #是否 key 一樣
  def is_same_key(self, word):
    return self.spell == word.spell and self.pron == word.pron




#处理Excel  
class ExcelHandler:
  
  def __init__(self, filename):
    self.filename=filename
  
  
  #获取Excel中单词的key
  def get_key(self, sheetname=None): 
    wb = openpyxl.load_workbook(self.filename) #打开Excel
    if sheetname is not None:
      sheet = wb[sheetname] #定位表单
    else:
      sheet = wb.active  # 获取当前活动工作表的对象
      
    keys = [] #创建一个空列表
    for row in range(2, sheet.max_row + 1):
      keys.append((sheet.cell(row, 1).value , sheet.cell(row, 2).value)) #将每行的数据循环加到列表中
    wb.close()  
    return keys


  #获取Excel中单词的data
  def get_data(self, sheetname=None): 
    wb = openpyxl.load_workbook(self.filename) #打开Excel
    if sheetname is not None:
      sheet = wb[sheetname] #定位表单
    else:
      sheet = wb.active  # 获取当前活动工作表的对象
      
    words = [] #创建一个空列表
    for row in range(2, sheet.max_row + 1):       #将每行的数据循环加到列表中  
      
      spell = sheet.cell(row, 1).value
      pron = sheet.cell(row, 2).value
      accent = sheet.cell(row, 3).value
      word_class = sheet.cell(row, 4).value
      meanings = sheet.cell(row, 5).value
      sentence = sheet.cell(row, 6).value
      translation = sheet.cell(row, 7).value
      word = Words(spell, pron, accent, word_class , meanings , sentence , translation)
      
      words.append(word) 
      
    wb.close()  
    return words
  
  
  
  """
  将制定的信息保存到新建的excel表格中;
  
  :param wbname:
  :param data: 往excel中存储的数据;
  :param sheetname:
  :return:
  """
  @staticmethod
  def create_to_excel(wbname, words, sheetname='Sheet1'):
  
      print("正在创建excel表格%s......" % (wbname))
  
      # 将数据data写入excel表格中;
      print("正在实例化excel表格%s......" % (wbname))
      #  如果文件不存在， 自己实例化一个WorkBook的对象;
      if(os.path.exists(wbname)):
        wb = openpyxl.load_workbook(wbname) #打开Excel
        if sheetname in wb.sheetnames:
          ws = wb[sheetname] #定位表单
        else:
          ws = wb.create_sheet(sheetname)        
      else:
        # Create a Workbook  
        wb = openpyxl.Workbook() 
        
        # 获取当前活动工作表的对象
        ws = wb.active
        #  Now change the name of Worksheet to “Changed Sheet” .
        ws.title = sheetname

      # 将数据data写入excel表格中;
      print("正在写入数据........")
      
      #首行       
      ws.cell(1, 1, "表記")
      ws.cell(1, 2, "読み")
      ws.cell(1, 3, "声调")
      ws.cell(1, 4, "词性")
      ws.cell(1, 5, "词义")
      ws.cell(1, 6, "文節")
      ws.cell(1, 7, "意味")
      
#       for row in range(0, len(words)):  # 第2行開始
      for word in words:  # 第2行開始
#         word = words[row]
        row = ws.max_row + 1  # 開始
  
        ws.cell(row, 1, word.spell)
        ws.cell(row, 2, word.pron)
        ws.cell(row, 3, word.accent)
        ws.cell(row, 4, word.word_class)
        ws.cell(row, 5, word.meanings)
        ws.cell(row, 6, word.sentence)
        ws.cell(row, 7, word.translation)
            
            
      # Save a file as sample_book.xlsx with save function.           
      wb.save(wbname)
      
      print("保存工作薄%s成功......." % (wbname))
      


if __name__ == "__main__":    
  
  list1 = ExcelHandler("D:/back/N1必背.xlsx").get_data("平仮名")  
  print(list1)
  
  list2 = ExcelHandler("D:/back/JLPT单词.xlsm").get_data("N1級")  
  print(list2)
  
  list3 = ExcelHandler("D:/back/JLPT单词.xlsm").get_data("N2級")  
  print(list3)
  
  list4 = ExcelHandler("D:/back/JLPT单词.xlsm").get_data("3、4級")  
  print(list4)
  
  list5 = ExcelHandler("D:/back/JLPT单词.xlsm").get_data("接头接尾")  
  print(list5)
  
  list2.extend(list3)
  list2.extend(list4)
  list2.extend(list5)
  
  samewords = [] #创建一个空列表
  samewords2 = [] #创建一个空列表
  # 列表的逆序遍历  remove
  for word1 in reversed(list1):     
#     print(word1.spell,word1.pron);
    for word2 in reversed(list2):     
#       if word1.is_same_key(word2):
      if str(word1.pron).strip() == str(word2.pron).strip():  

        if word1 not in samewords:
          samewords.append(word1)       
        if word2 not in samewords2:
          samewords2.append(word2)
        
        if word2 in list2:
          list2.remove(word2)
          
        if word1 in list1:
          list1.remove(word1)
#         break
      
  ExcelHandler.create_to_excel('D:/back/hello.xlsx', list1,"平仮名")
  ExcelHandler.create_to_excel("D:/back/hello.xlsx", reversed(samewords), "重なる1")
  ExcelHandler.create_to_excel("D:/back/hello.xlsx", reversed(samewords2), "重なる2")
  
  