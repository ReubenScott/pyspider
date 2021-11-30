#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import re
import os
import time
import random
import openpyxl
import traceback
import requests
from requests.adapters import HTTPAdapter
from bs4 import BeautifulSoup, SoupStrainer

#代理池
proxies = [
  "http://190.186.1.46:55830",
  "http://62.205.169.74:53281"
]


#模擬header的user-agent字段，返回一個隨機的user-agent字典類型的鍵值對
useragents = [
  'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0',  
  'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.100 Safari/537.36',
  'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
  'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36 Edge/17.17134',
  'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36 Edg/90.0.818.51',
  'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0'
]

# 配置请求头
headers = {
  'Accept-Encoding': 'gzip, deflate, br',
	'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
  'Content-Type': 'text/html; charset=utf-8',
  'Connection': 'close'
}

cookies = [
    'HJ_UID=1d97c1e3-4982-851d-e2fb-9688a9cf8765; TRACKSITEMAP=3%2C6%2C20%2C23; _REF=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%25E8%25B2%25B4%25E6%2596%25B9; _REG=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_3=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%E8%B2%B4%E6%96%B9; _SREG_3=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_20=https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DhuEIcpnuOcb6ULzhdTnwE8xiX4uRF1EAwlFc8cPh8rszH-a6knkQEEWAd0HYEnmT%26wd%3D%26eqid%3Df653c792000568c5000000056072c5e8; _SREG_20=www.baidu.com%7C%7Csearch%7Cdomain; _UZT_USER_SET_106_0_DEFAULT=2|6ed2e0a3900e22cb93d3dd68d679fad2; HJ_SID=djz0xw-d350-411a-be29-4bbf008627f8; HJ_SSID_3=djz0xw-935d-439d-be26-7ea49f256106; HJ_CST=0; HJ_CSST_3=1; acw_tc=2f624a2416189375318547181e14f93a3a5e54a9bf667711cfa1f39838005a',
    'HJ_UID=61fb377b-88ee-a3a1-8058-718ab5f6097d; TRACKSITEMAP=3%2C20%2C23; _REF=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%E3%81%82%E3%81%A1%E3%82%89; _REG=direct%7C%7Cdirect%7Cdirect; _SREF_3=; _SREG_3=direct%7C%7Cdirect%7Cdirect; _SREF_20=https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DKCJVBw943ECLx9-ap81lRDx4DJ2wdr4MMK9-HY1KCPuYYw6vkVwt1oeBc50rSX41%26wd%3D%26eqid%3Dec26d1f80004eeeb00000004607d1807; _SREG_20=www.baidu.com%7C%7Csearch%7Cdomain; Hm_lvt_d4f3d19993ee3fa579a64f42d860c2a7=1617957780,1618021323,1618294620,1618810892; _UZT_USER_SET_106_0_DEFAULT=2|6ed2e0a3900e22cb93d3dd68d679fad2; Hm_lpvt_d4f3d19993ee3fa579a64f42d860c2a7=1618810892; HJ_SID=oxl2bb-51af-4dfc-bf89-17bbfbdd60b2; HJ_SSID_3=oxl2bb-7b2e-4483-8bee-ee21f321ccb5; HJ_CST=0; HJ_CSST_3=1; acw_tc=707c9f9d16189714475066449e01f4914d4d6a1dfe79bfac97268fb6681cd4',
    'HJ_UID=61fb377b-88ee-a3a1-8058-718ab5f6097d; TRACKSITEMAP=3%2C20%2C23; _REF=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%25E7%25BE%258A%25E6%25AF%259B; _REG=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_3=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%E7%BE%8A%E6%AF%9B; _SREG_3=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_20=https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DKCJVBw943ECLx9-ap81lRDx4DJ2wdr4MMK9-HY1KCPuYYw6vkVwt1oeBc50rSX41%26wd%3D%26eqid%3Dec26d1f80004eeeb00000004607d1807; _SREG_20=www.baidu.com%7C%7Csearch%7Cdomain; Hm_lvt_d4f3d19993ee3fa579a64f42d860c2a7=1617957780,1618021323,1618294620,1618810892; _UZT_USER_SET_106_0_DEFAULT=2|6ed2e0a3900e22cb93d3dd68d679fad2; Hm_lpvt_d4f3d19993ee3fa579a64f42d860c2a7=1618810892; acw_tc=2f624a4816189053457392237e0df8c605cf50e3a5b2d389c20027d1f163b1; HJ_SID=ra4zlx-1e1f-4399-b195-c4b0dd46e3c0; HJ_SSID_3=ra4zlx-5ee0-43a1-b8d9-269d60aa68a8; HJ_CST=0; HJ_CSST_3=1'
  ]


#单词
class Words:
                    
  #表記  英式  美式  词性  词义  文節（詞組）  意味
  def __init__(self, spell, pron_en, pron_us, word_class , meanings , sentence , translation):
    self.spell = spell
    self.pron_en = pron_en
    self.pron_us = pron_us
    self.word_class = word_class
    self.meanings = meanings
    self.sentence = sentence
    self.translation = translation

  def __eq__(self, other):
    if not isinstance(other, Words):
      return NotImplemented
    else:
      # string lists of all method names and properties of each of these objects
      prop_names1 = list(self.__dict__)
      prop_names2 = list(other.__dict__)

      n = len(prop_names1)  # number of properties
      for i in range(n):
        if getattr(self, prop_names1[i]) != getattr(other, prop_names2[i]):
          return False

      return True

  def __hash__(self):
    # necessary for instances to behave sanely in dicts and sets.
    return hash((self.foo, self.bar))
  
  def __str__(self):
    return "%s | %s | %s | %s | %s | %s | %s " %(self.spell, self.pron_en, self.pron_us, self.word_class , self.meanings , self.sentence , self.translation)

  #是否要更新  
  def need_fix(self):
    return not (self.spell and self.pron_en and self.pron_us and self.word_class and self.meanings and self.sentence and self.translation)


  #是否 key 一樣
  def is_same_key(self, word):
    return self.spell == word.spell and self.pron_en == word.pron_en




#处理Excel  
class ExcelHandler:
  
  def __init__(self, filename):
    self.filename=filename
  
  
  #获取Excel中单词的key
  def get_key(self, sheetname=None): 
    wb = openpyxl.load_workbook(self.filename) #打开Excel
    if sheetname is not None:
      sheet = wb[sheetname] #定位表单
    else:
      sheet = wb.active  # 获取当前活动工作表的对象
      
    keys = [] #创建一个空列表
    for row in range(2, sheet.max_row + 1):
      keys.append((sheet.cell(row, 1).value , sheet.cell(row, 2).value)) #将每行的数据循环加到列表中
    wb.close()  
    return keys


  #获取Excel中单词的data
  def get_data(self, sheetname=None): 
    wb = openpyxl.load_workbook(self.filename) #打开Excel
    if sheetname is not None:
      sheet = wb[sheetname] #定位表单
    else:
      sheet = wb.active  # 获取当前活动工作表的对象
      
    words = [] #创建一个空列表
    for row in range(2, sheet.max_row + 1):       #将每行的数据循环加到列表中  
      
      spell = sheet.cell(row, 1).value
      pron_en = sheet.cell(row, 2).value
      pron_us = sheet.cell(row, 3).value
      word_class = sheet.cell(row, 4).value
      meanings = sheet.cell(row, 5).value
      sentence = sheet.cell(row, 6).value
      translation = sheet.cell(row, 7).value
      word = Words(spell, pron_en, pron_us, word_class , meanings , sentence , translation)
      
      words.append(word) 
      
    wb.close()  
    return words
  
  
  
  """
  将制定的信息保存到新建的excel表格中;
  
  :param wbname:
  :param data: 往excel中存储的数据;
  :param sheetname:
  :return:
  """
  @staticmethod
  def create_to_excel(wbname, words, sheetname='Sheet1'):
  
      print("正在创建excel表格%s......" % (wbname))
  
      # 将数据data写入excel表格中;
      print("正在实例化excel表格%s......" % (wbname))
      #  如果文件不存在， 自己实例化一个WorkBook的对象;
      if(os.path.exists(wbname)):
        wb = openpyxl.load_workbook(wbname) #打开Excel
        if sheetname in wb.sheetnames:
          ws = wb[sheetname] #定位表单
        else:
          ws = wb.create_sheet(sheetname)        
      else:
        # Create a Workbook  
        wb = openpyxl.Workbook() 
        
        # 获取当前活动工作表的对象
        ws = wb.active
        #  Now change the name of Worksheet to “Changed Sheet” .
        ws.title = sheetname

      # 将数据data写入excel表格中;
      print("正在写入数据........")
      
      #首行       
      ws.cell(1, 1, "表記")
      ws.cell(1, 2, "英式")  
      ws.cell(1, 3, "美式")
      ws.cell(1, 4, "词性")
      ws.cell(1, 5, "词义")
      ws.cell(1, 6, "文節")
      ws.cell(1, 7, "意味")
      
#       for row in range(0, len(words)):  # 第2行開始
      for word in words:  # 第2行開始
#         word = words[row]
        row = ws.max_row + 1  # 開始
  
        ws.cell(row, 1, word.spell)
        ws.cell(row, 2, word.pron_en)
        ws.cell(row, 3, word.pron_us)
        ws.cell(row, 4, word.word_class)
        ws.cell(row, 5, word.meanings)
        ws.cell(row, 6, word.sentence)
        ws.cell(row, 7, word.translation)
            
            
      # Save a file as sample_book.xlsx with save function.           
      wb.save(wbname)
      
      print("保存工作薄%s成功......." % (wbname))



# 超時時間(秒)
TIMEOUT = 10
# 重连次数
MAX_RETRIES = 3

# get获取請求数据
def get_content(session, url):
  resp = None
  try:
    headers['User-Agent'] = useragents[random.randint(0, len(useragents)-1)]
    headers['Cookie'] = cookies[random.randint(0, len(cookies)-1)]
    resp = session.get(url, headers=headers, timeout=TIMEOUT)
#     resp = session.get(url, headers=headers, proxies={'http': proxies[random.randint(0, len(proxies)-1)]}, timeout=TIMEOUT)
    resp.raise_for_status()   # 如果返回的状态码不是200， 则抛出异常;
    resp.encoding = resp.apparent_encoding  # 判断网页的编码格式， 便于respons.text知道如何解码;
  except:
    print(traceback.format_exc())
  else:
    return resp
  finally:
    # 注意关闭response 
    if resp:
      resp.close()  
    

# post获取請求数据
def post_content(session, url, data):
  resp = None
  try:
    headers['User-Agent'] = useragents[random.randint(0, len(useragents)-1)]
    headers['Cookie'] = cookies[random.randint(0, len(cookies)-1)]
    resp = session.post(url, data=data, headers=headers, timeout=TIMEOUT)
    resp.raise_for_status()   # 如果返回的状态码不是200， 则抛出异常;
    resp.encoding = resp.apparent_encoding  # 判断网页的编码格式， 便于respons.text知道如何解码;
  except:
    print(traceback.format_exc())
  else:
    return resp
  finally:
    # 注意关闭response 
    if resp:
      resp.close()  


# 查詢單詞  https://dict.hjenglish.com
def search_hjclass_dict(session, word):
  searchText = word.spell
  
  if searchText:
    url = 'https://dict.hjenglish.com/w/' + searchText
        
    resp = get_content(session, url)
    # 获取状态 
#     print(url, resp.status_code)
#     print(response.content.decode())

    # 创建soup对象 
    soup = BeautifulSoup(resp.text, "html.parser") 
#     example = soup.find('div', {'class': 'word-details '})

    # 多種讀音
    if len(soup.select('div.word-details div.word-details-pane')) > 0 : 
      soup = BeautifulSoup(soup.select('div.word-details div.word-details-pane')[0].prettify(), "html.parser")    
     
    #單詞
    if len(soup.select('div.word-info div.word-text h2')) > 0 :
      spell = soup.select('div.word-info div.word-text h2')[0].text.strip()
    
      if spell == searchText: 
        #發音
        pronounces = soup.select('div.pronounces span')
        if len(pronounces) == 6 : 
          word.pron_en = '/' + pronounces[1].text.strip()[1:-1] + '/' # 音標 [ˈpʌpɪ]
          word.accent_en = pronounces[2].attrs['data-src']  #  word-audio-en data-src="https://tts.hjapi.com/en-gb/61A364FE49D45BBB"
        
          word.pron_us = '/' + pronounces[4].text.strip()[1:-1] + '/'  # 音標 [ˈpʌpɪ]
          word.accent_us = pronounces[5].attrs['data-src'] 
        
        #詞性
        if len(soup.select('div.simple span')) > 0 :
          word.word_class = soup.select('div.simple span')[0].text.strip()[0:-1]
          #詞义
          word.meanings = soup.select('div.simple span')[1].text.strip()

        #例句
        example_sentence = soup.select('div.word-details-item-content section.detail-groups p.def-sentence-from')
        word.sentence = (example_sentence[0].text.strip() if (len(example_sentence) > 0) else '')
        
        example_translation = soup.select('div.word-details-item-content section.detail-groups p.def-sentence-to')
        word.translation = (example_translation[0].text.strip() if (len(example_translation) > 0) else '')



if __name__ == "__main__":    
  
  words = ExcelHandler("D:/Back/英语词汇表.xlsx").get_data("Sheet1")  
  print(words)

  session = requests.session()   
  session.keep_alive = False # 设置连接活跃状态为False
  #设置重连次数
  session.mount('http://', HTTPAdapter(max_retries=MAX_RETRIES))
  session.mount('https://', HTTPAdapter(max_retries=MAX_RETRIES))
  
  
  # 爬取要素
  for word in words:
    if word.need_fix():
      print(word)
      search_hjclass_dict(session, word)
      print(word)
      
  ExcelHandler.create_to_excel('D:/back/hello.xlsx', words)

  