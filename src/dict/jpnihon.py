#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import openpyxl
import Words

#处理Excel
class doExcel:
  
  def __init__(self,file_name,sheet_name=""):
    self.filename=file_name
    self.sheetname=sheet_name
  
  
  #获取Excel中的方法
  def get_key(self): 
    wb = openpyxl.load_workbook(self.filename, data_only=True) #打开Excel
#     ⓪　① ② ③ ④ ⑤ ⑥ ⑦ ⑧ ⑨
    pron_flag = ('⓪','①','②','③','④','⑤','⑥','⑦','⑧','⑨')
    
    index = 0
    words = [] #创建一个空列表
    keys = [] #创建一个空列表
    
    add_flag = ""
    spell, pron, accent, word_class , meanings , sentence , translation = "","","","","","",""
    
    for sheet in wb:
      sheet = wb[sheet.title] #定位表单
      
      # 获得指定行列的单元格      for row in sheet.rows:
      for row in sheet.iter_rows(max_row=sheet.max_row -1 , min_row=1):
         
        line = ""      
        # 获取单元格的值
        for cell in row:
          if cell.value:
            line = line + cell.value.strip()
            
        for flag in pron_flag:
          if (flag in line):
            index = 1
            spell = line
            continue
        
        if (add_flag != "") and (spell != add_flag) :
          word = Words.Words(add_flag, pron, accent, word_class , meanings , sentence , translation)
          words.append(word) 
          pron, accent, word_class , meanings , sentence , translation = "","","","","",""
          
        if index >= 3:
          sentence =  sentence + line
        
        if index == 2:
          word_class = line 
          
        if index == 1:
          add_flag = spell
          keys.append(spell)
          
        index =  index + 1
    
    wb.close()  
    
    return words
    
  
  
def create_to_excel(wbname, words, sheetname='Sheet1', ):
    """
    将制定的信息保存到新建的excel表格中;

    :param wbname:
    :param data: 往excel中存储的数据;
    :param sheetname:
    :return:
    """

    print("正在创建excel表格%s......" % (wbname))

    # wb = openpyxl.load_workbook(wbname)
    #  如果文件不存在， 自己实例化一个WorkBook的对象;
    wb = openpyxl.Workbook()
    # 获取当前活动工作表的对象
    sheet = wb.active
    # 修改工作表的名称
    sheet.title = sheetname
    # 将数据data写入excel表格中;
    print("正在写入数据........")
    
    for row in range(0, len(words)):  # 第2行開始
      word = words[row]
      row = sheet.max_row + 1  # 開始

      sheet.cell(row, 1, word.spell)
      sheet.cell(row, 2, word.pron)
      sheet.cell(row, 3, word.accent)
      sheet.cell(row, 4, word.word_class)
      sheet.cell(row, 5, word.meanings)
      sheet.cell(row, 6, word.sentence)
      sheet.cell(row, 7, word.translation)
          

    wb.save(wbname)
    print("保存工作薄%s成功......." % (wbname))



if __name__ == "__main__":    
  
  
  list1=doExcel("D:/N1必背2000词.xlsx").get_key()  
  
  
  create_to_excel('d:/hello.xlsx', list1)
#   set1 = set(list1)
#   print(len(set1))
#   set1 = set1 | set(list1)
  
#   print(set(list1) - set(list2))  # {2, 6}  
#   print(set2 - set1)  # {2, 6}  
  
         
  '''
  list2=doExcel("D:/temp/JLPT单词.xlsm","N1級").get_data()
  
  print(list2)  
  list3 = [item for item in list1 if item not in set(list2)]
  print(list3)  # [2, 6]
  '''
  
  
  
  