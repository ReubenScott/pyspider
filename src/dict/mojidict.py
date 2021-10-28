#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import json
import re
import os
import time
import openpyxl
from bs4 import BeautifulSoup, SoupStrainer

# 配置请求头
headers = {
  'User-Agent': 'Mozilla /5.0 (Macintosh; Intel Mac OS X 10_12_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
  'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,application/json,image/webp,image/apng ,*/*;q=0.8,application/signed-exchange;v=b3',
  'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
  'Accept-Encoding': 'gzip, deflate, br',
  'Content-Type': 'application/json;charset=utf-8',
  'Connection': 'close'
}

# 获取請求数据
def get_content(url):
  response = None
  while True:     
    try:
      response = requests.get(url, headers=headers, timeout=30)       
      response.raise_for_status()   # 如果返回的状态码不是200， 则抛出异常;
      response.encoding = response.apparent_encoding  # 判断网页的编码格式， 便于respons.text知道如何解码;
    except Exception as ex:
      time.sleep( 0.5)
      print(ex)
    else:
      return response.content



def create_to_excel(wbname, data, sheetname='Sheet1', ):
    """
    将制定的信息保存到新建的excel表格中;

    :param wbname:
    :param data: 往excel中存储的数据;
    :param sheetname:
    :return:
    """

    print("正在创建excel表格%s......" % (wbname))

    # wb = openpyxl.load_workbook(wbname)
    #  如果文件不存在， 自己实例化一个WorkBook的对象;
    wb = openpyxl.Workbook()
    # 获取当前活动工作表的对象
    sheet = wb.active
    # 修改工作表的名称
    sheet.title = sheetname
    # 将数据data写入excel表格中;
    print("正在写入数据........")
    for row, item in enumerate(data):  # data发现有4行数据， item里面有三列数据;
        print(item)
        for column, cellValue in enumerate(item):
            # cell = sheet.cell(row=row + 1, column=column + 1, value=cellValue)
            cell = sheet.cell(row=row+1, column=column + 1)
            cell.value = cellValue

    wb.save(wbname)
    print("保存工作薄%s成功......." % (wbname))



def search_moji_dict(searchText):
  if searchText:
    url = 'https://api.mojidict.com/parse/functions/search_v3'
    payload =   {
        "searchText": searchText,
        "needWords": 'true',
        "langEnv": "zh-CN_ja",
        "_ApplicationId": "E62VyFVLMiW7kvbtVq3p"
    }
    r = requests.post(url, headers=headers, timeout=30, data=json.dumps(payload))
    print(searchText , r.text)
    result = json.loads(r.text);
    print(result["result"]["words"])
    
    if len(result["result"]["words"]) > 0 :     
      searchText = result["result"]["searchResults"][0]["searchText"]
      spell = result["result"]["words"][0]["spell"]
      
      if spell in searchText: 
        pron = result["result"]["words"][0]["pron"]
        if "accent" in result["result"]["words"][0]:
          accent = result["result"]["words"][0]["accent"]
        else:
          accent = ''
        if "excerpt" in result["result"]["words"][0]:
          excerpt = result["result"]["words"][0]["excerpt"]
        else:
          excerpt = ''
        # objectId  https://www.mojidict.com/details/198963336
        objectId = result["result"]["words"][0]["objectId"]
          
        #獲取
        pattern = re.compile("(^\[.*?\])(.*)", re.MULTILINE | re.DOTALL)     
        # pattern.search(excerpt).groups()
        picker = pattern.search(excerpt)
        if picker:
          picker = picker.groups()
          print(picker[0])
          print(picker[1])
          
          # 创建soup对象 
          soup = BeautifulSoup(get_content('https://www.mojidict.com/details/' + objectId), "html.parser") 
          example = soup.find('div', {'class': 'example-info'})
          print(example)
      
          if example:
            example_item = []
            for i,child in enumerate(example.contents):
              if hasattr(child, 'div'):
                print(i, child.string.strip())
                example_item.append(child.string.strip())
          
            return (spell, pron, accent, picker[0] , picker[1]) + tuple(example_item)


#处理Excel
class doExcel:
  def __init__(self,file_name,sheet_name):
    self.filename=file_name
    self.sheetname=sheet_name
      
  #获取Excel中的方法
  def get_data(self): 
    wb = openpyxl.load_workbook(self.filename) #打开Excel
    sheet = wb[self.sheetname] #定位表单
    test_data= [] #创建一个空列表
    for row in range(2, sheet.max_row + 1):
      test_data.append(sheet.cell(row, 1).value) #将每行的数据循环加到列表中
    wb.close()  
    return test_data




#处理Excel
class Words:
                    
   #表記  読み  声调  词性  词义  文節（詞組）  意味
   def __init__(self, spell, pron, accent, word_class , meanings , sentence , translation):
      self.spell = spell
      self.pron = pron
      self.accent = accent
      self.word_class = word_class
      self.meanings = meanings
      self.sentence = sentence
      self.translation = translation




if __name__ == "__main__":    
  
  res=doExcel("F:/OneDrive/Documents/单词/新标日初级中级单词表.xlsx","中级单词").get_data()
  print(res)
  
  movieInfo = []
  
  for word in res:
    print(word)
    item = search_moji_dict(word)
    if item is not None:
      movieInfo.append(item)
#     count = count + 1
#     if count == 20 :
#       break;
    
  print(movieInfo)  
  create_to_excel('d:/hello.xlsx', movieInfo)


