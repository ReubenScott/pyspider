#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import json
import re
import os
import time
import random
import openpyxl
from bs4 import BeautifulSoup, SoupStrainer

#代理池
proxies = [
  "http://190.186.1.46:55830",
  "http://1.20.99.89:32963",
  "http://149.129.62.207:3128",
  "http://47.75.90.57:80",
  "http://51.91.157.66:80",
  "http://91.209.11.131:80",
  "http://150.109.32.166:80",
  "http://69.167.174.17:80",
  "http://70.35.205.79:80",
  "http://91.92.180.45:8080",
  "http://136.243.211.104:80",
  "http://200.35.56.161:35945",
  "http://93.188.161.84:80",
  "http://95.79.55.196:53281",
  "http://195.138.73.54:44017",
  "http://188.68.52.45:80",
  "http://110.74.208.154:21776",
  "http://188.166.204.196:80",
  "http://115.85.65.147:39736",
  "http://67.205.135.215:80",
  "http://220.135.165.38:8080",
  "http://91.209.11.132:80",
  "http://95.161.188.246:38302",
  "http://68.183.88.73:80",
  "http://202.70.67.93:80",
  "http://218.253.39.60:8383",
  "http://77.68.125.33:80",
  "http://165.22.252.119:80",
  "http://193.34.55.64:32767",
  "http://109.50.177.30:80",
  "http://167.99.174.59:80",
  "http://81.174.23.244:80",
  "http://52.140.103.185:80",
  "http://77.68.29.157:80",
  "http://190.112.136.185:8085",
  "http://132.148.85.91:80",
  "http://159.65.69.186:9300",
  "http://158.140.167.148:53281",
  "http://80.26.96.212:80",
  "http://138.197.102.119:80",
  "http://181.129.70.82:46752",
  "http://119.2.47.46:8080",
  "http://119.82.252.25:42914",
  "http://136.232.209.70:47423",
  "http://142.93.223.219:8080",
  "http://79.150.195.72:80",
  "http://74.205.128.201:80",
  "http://85.214.83.135:8085",
  "http://200.29.148.178:8080",
  "http://62.205.169.74:53281"
]


#模擬header的user-agent字段，返回一個隨機的user-agent字典類型的鍵值對
useragents = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0',  
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.100 Safari/537.36',
            'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
            'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36 Edge/17.17134',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36 Edg/90.0.818.51',
            'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0'
           ]

# 配置请求头
fakeheaders = {
  'Accept-Encoding': 'gzip, deflate, br',
	'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
  'Content-Type': 'text/html; charset=utf-8',
  'Connection': 'close'
}

cookies = [
    'HJ_UID=1d97c1e3-4982-851d-e2fb-9688a9cf8765; TRACKSITEMAP=3%2C6%2C20%2C23; _REF=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%25E8%25B2%25B4%25E6%2596%25B9; _REG=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_3=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%E8%B2%B4%E6%96%B9; _SREG_3=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_20=https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DhuEIcpnuOcb6ULzhdTnwE8xiX4uRF1EAwlFc8cPh8rszH-a6knkQEEWAd0HYEnmT%26wd%3D%26eqid%3Df653c792000568c5000000056072c5e8; _SREG_20=www.baidu.com%7C%7Csearch%7Cdomain; _UZT_USER_SET_106_0_DEFAULT=2|6ed2e0a3900e22cb93d3dd68d679fad2; HJ_SID=djz0xw-d350-411a-be29-4bbf008627f8; HJ_SSID_3=djz0xw-935d-439d-be26-7ea49f256106; HJ_CST=0; HJ_CSST_3=1; acw_tc=2f624a2416189375318547181e14f93a3a5e54a9bf667711cfa1f39838005a',
    'HJ_UID=61fb377b-88ee-a3a1-8058-718ab5f6097d; TRACKSITEMAP=3%2C20%2C23; _REF=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%E3%81%82%E3%81%A1%E3%82%89; _REG=direct%7C%7Cdirect%7Cdirect; _SREF_3=; _SREG_3=direct%7C%7Cdirect%7Cdirect; _SREF_20=https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DKCJVBw943ECLx9-ap81lRDx4DJ2wdr4MMK9-HY1KCPuYYw6vkVwt1oeBc50rSX41%26wd%3D%26eqid%3Dec26d1f80004eeeb00000004607d1807; _SREG_20=www.baidu.com%7C%7Csearch%7Cdomain; Hm_lvt_d4f3d19993ee3fa579a64f42d860c2a7=1617957780,1618021323,1618294620,1618810892; _UZT_USER_SET_106_0_DEFAULT=2|6ed2e0a3900e22cb93d3dd68d679fad2; Hm_lpvt_d4f3d19993ee3fa579a64f42d860c2a7=1618810892; HJ_SID=oxl2bb-51af-4dfc-bf89-17bbfbdd60b2; HJ_SSID_3=oxl2bb-7b2e-4483-8bee-ee21f321ccb5; HJ_CST=0; HJ_CSST_3=1; acw_tc=707c9f9d16189714475066449e01f4914d4d6a1dfe79bfac97268fb6681cd4',
    'HJ_UID=61fb377b-88ee-a3a1-8058-718ab5f6097d; TRACKSITEMAP=3%2C20%2C23; _REF=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%25E7%25BE%258A%25E6%25AF%259B; _REG=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_3=https%3A%2F%2Fdict.hjenglish.com%2Fjp%2Fjc%2F%E7%BE%8A%E6%AF%9B; _SREG_3=dict.hjenglish.com%7C%7Cxiaodi_site%7Cdomain; _SREF_20=https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DKCJVBw943ECLx9-ap81lRDx4DJ2wdr4MMK9-HY1KCPuYYw6vkVwt1oeBc50rSX41%26wd%3D%26eqid%3Dec26d1f80004eeeb00000004607d1807; _SREG_20=www.baidu.com%7C%7Csearch%7Cdomain; Hm_lvt_d4f3d19993ee3fa579a64f42d860c2a7=1617957780,1618021323,1618294620,1618810892; _UZT_USER_SET_106_0_DEFAULT=2|6ed2e0a3900e22cb93d3dd68d679fad2; Hm_lpvt_d4f3d19993ee3fa579a64f42d860c2a7=1618810892; acw_tc=2f624a4816189053457392237e0df8c605cf50e3a5b2d389c20027d1f163b1; HJ_SID=ra4zlx-1e1f-4399-b195-c4b0dd46e3c0; HJ_SSID_3=ra4zlx-5ee0-43a1-b8d9-269d60aa68a8; HJ_CST=0; HJ_CSST_3=1'
  ]

 

# 获取請求数据
def search_hj_dict(searchText , pronounce):
  time.sleep(5) #設置查詢間隔
  if searchText:
    fakeheaders['User-Agent'] = useragents[random.randint(0, len(useragents)-1)]
    fakeheaders['Cookie'] = cookies[random.randint(0, len(cookies)-1)]
    url = 'https://dict.hjenglish.com/jp/jc/' + searchText
        
    session = requests.session()
    resp = session.get(url, headers=fakeheaders, proxies={'http': proxies[random.randint(0, len(proxies)-1)]}, timeout=30)
#     resp = requests.get(url, headers=fakeheaders, proxies={'http': proxies[random.randint(0, len(proxies)-1)]}, timeout=30)
    # 获取状态 
    print(url, resp.status_code)
#     print(response.content.decode())
    # 创建soup对象 
    session.close()
    soup = BeautifulSoup(resp.text, "html.parser") 
#     example = soup.find('div', {'class': 'word-details '})

    # 多種讀音
    if len(soup.select('header.word-details-header ul div.pronounces span.pronounce-value')) > 0 : 
      for item in soup.select('section.word-details-content div.word-details-pane'): 
        soup = BeautifulSoup(item.decode(), "html.parser") 
        #字符串截取
        if soup.select('div.word-info div.pronounces span')[0].text.strip()[1:-1] in pronounce: 
          break;
    elif len(soup.select('div.word-details div.word-details-pane')) > 0 : 
      soup = BeautifulSoup(soup.select('div.word-details div.word-details-pane')[0].prettify(), "html.parser")    
     
    #單詞
    if len(soup.select('div.word-info div.word-text h2')) > 0 :
      spell = soup.select('div.word-info div.word-text h2')[0].text.strip()
    
      if spell in searchText: 
        #發音
        pronounces = soup.select('div.pronounces span')
        if len(pronounces) == 4 : 
          pron = pronounces[0].text.strip()[1:-1] #字符串截取
          print(pron)
          print(pronounces[1].text.strip()[1:-1])
          accent = pronounces[2].text.strip()
          print(pronounces[3].attrs['data-src'])
        
          #詞性
          if len(soup.select('div.simple h2')) > 0 :
            word_class = soup.select('div.simple h2')[0].text.strip()[1:-1]
            
            #詞义
            meanings = soup.select('div.simple ul li')
            word_meanings = ''
            for meaning in meanings:
              word_meanings = word_meanings + ";".join(meaning.text.split()) 
            
            #例句
            example_sentence = soup.select('div.word-details-item-content section.detail-groups p.def-sentence-from')
            example_sentence = (example_sentence[0].text.strip() if (len(example_sentence) > 0) else '')
            
            example_translation = soup.select('div.word-details-item-content section.detail-groups p.def-sentence-to')
            example_translation = (example_translation[0].text.strip() if (len(example_translation) > 0) else '')
            
            print(spell, pron, accent, word_class , word_meanings , example_sentence , example_translation)
            # 汉字  假名(读音) 声调   词性   词义   文節（詞組）   意味
            return (spell, pron, accent, word_class , word_meanings , example_sentence , example_translation)


#处理Excel
def save_to_excel(wbname, sheetname='Sheet1' ):
    """
         将制定的信息保存到新建的excel表格中;
    :param wbname:
    :param data: 往excel中存储的数据;
    :param sheetname:
    :return:
    """

    print("正在实例化excel表格%s......" % (wbname))
    # 打开Excel， 实例化一个WorkBook的对象;
    wb = openpyxl.load_workbook(wbname) 
    # 获取当前活动工作表的对象
    # sheet = wb.active
    sheet = wb[sheetname] #定位表单
    # 将数据data写入excel表格中;
    print("正在写入数据........")
    for row in range(2, sheet.max_row + 1):  # 第2行開始
      accent = sheet.cell(row, 3).value
      if not accent:
        content = search_hj_dict(sheet.cell(row, 1).value, sheet.cell(row, 2).value)
        # print(row , content[2],  content[5],  content[6])
        # cell = sheet.cell(row=row + 1, column=column + 1, value=cellValue)
        if content and content[1] in sheet.cell(row, 2).value :
          cell = sheet.cell(row, 3)
          cell.value = content[2]
          if not sheet.cell(row, 4).value:
            sheet.cell(row, 4, content[3])
          if len(content) > 5:
            sheet.cell(row=row, column=6, value=content[5])
          if len(content) > 6:
            sheet.cell(row, 7, content[6])
          break  
      else:
        continue
      
    wb.save(wbname)
    print("保存工作薄%s成功......." % (wbname))

if __name__ == "__main__":    
  save_to_excel("D:/日语N2单词表.xlsx","N2級单词表")
  
  